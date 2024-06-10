# -*- mode:python; coding:utf-8 -*-
# Copyright (c) 2024 IBM Corp. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""AI_RMF catalog creator."""

import argparse
import datetime
import os
import sys
import uuid
from pathlib import Path
from typing import Dict, Iterator

from openpyxl import load_workbook

from trestle.oscal import OSCAL_VERSION
from trestle.oscal.catalog import Catalog, Control, Group
from trestle.oscal.common import Metadata, Property

import yaml

timestamp = (datetime.datetime.utcnow().replace(microsecond=0).replace(tzinfo=datetime.timezone.utc).isoformat())


def info(text: str) -> None:
    """Issue info message."""
    print(f'I {text}')


class XlsxHelper:
    """Xlsx Helper common functions and assistance navigating spread sheet."""

    def __init__(self, file: str, sheet_name: str) -> None:
        """Initialize."""
        self._spread_sheet = file
        self._wb = load_workbook(self._spread_sheet)
        self._sheet_name = sheet_name
        self._work_sheet = self._wb[self._sheet_name]
        self._mapper()

    def _normalize(self, name: str) -> str:
        """Normalize."""
        return name.lower()

    def _translate(self, name: str) -> str:
        """Translate name key to column name."""
        return name

    def _mapper(self) -> None:
        """Map columns heading names to column numbers."""
        self._col_name_to_number = {}
        cols = self._work_sheet.max_column + 1
        row = 1
        for col in range(1, cols):
            cell = self._work_sheet.cell(row, col)
            if cell.value:
                name = self._normalize(cell.value)
                self._col_name_to_number[name] = col

    def row_generator(self) -> Iterator[int]:
        """Generate rows until max reached."""
        row = 2
        while row <= self._work_sheet.max_row:
            yield row
            row += 1

    def _get(self, row: int, name: str) -> str:
        """Get cell value for given row and column name."""
        nname = self._normalize(name)
        cname = self._translate(nname)
        col = self._col_name_to_number[cname]
        cell = self._work_sheet.cell(row, col)
        return cell.value

    def get_group_id(self, row: int) -> str:
        """Get group id."""
        return self._get(row, 'Group Id')

    def get_group_title(self, row: int) -> str:
        """Get group title."""
        return self._get(row, 'Group Title')

    def get_subgroup_id(self, row: int) -> str:
        """Get subgroup id."""
        return self._get(row, 'Sub Group Id')

    def get_subgroup_title(self, row: int) -> str:
        """Get subgroup title."""
        return self._get(row, 'Sub Group Title')

    def get_control_id(self, row: int) -> str:
        """Get control id."""
        return self._get(row, 'Control Id')

    def get_control_title(self, row: int) -> str:
        """Get control_title."""
        return self._get(row, 'Control Title')

    def get_control_description(self, row: int) -> str:
        """Get control_description."""
        return self._get(row, 'Control Description')


class CatalogHelper:
    """OSCAL Catalog Helper."""

    def __init__(self, title: str, version: str) -> None:
        """Initialize."""
        self.groups = []
        # metadata
        self._metadata = Metadata(
            title=title,
            last_modified=timestamp,
            oscal_version=OSCAL_VERSION,
            version=version,
        )

    def get_catalog(self) -> Catalog:
        """Get catalog."""
        catalog = Catalog(
            uuid=str(uuid.uuid4()),
            metadata=self._metadata,
        )
        catalog.groups = self.groups
        return catalog

    def _get_group(self, groups, id_, title, subgroups, controls) -> Group:
        """Get group."""
        rval = None
        for group in groups:
            if group.id == id_:
                rval = group
                break
        if not rval:
            group = Group(id=id_, title=title, groups=subgroups, controls=controls)
            groups.append(group)
            rval = group
        return rval

    def _normalize(self, text) -> str:
        """Normalize."""
        value = text
        if value:
            value = value.strip()
            value = value.split()
            value = ' '.join(value)
        return value

    def add_control(
        self,
        group_id: str,
        group_title: str,
        subgroup_id: str,
        subgroup_title: str,
        control_id: str,
        control_title,
        control_description: str,
    ) -> None:
        """Add control."""
        group = self._get_group(self.groups, group_id, group_title, [], None)
        subgroup = self._get_group(group.groups, subgroup_id, subgroup_title, None, [])
        control = Control(id=control_id, title=control_title)
        subgroup.controls.append(control)
        control.props = []
        value = self._normalize(control_description)
        if value:
            prop = Property(name='Control_Description', value=value)
            control.props.append(prop)


class CatalogBuilder:
    """OSCAL catalog builder."""

    def __init__(self) -> None:
        """Initialize."""

    def run(self) -> None:
        """Run."""
        # parse
        args = self.parse_args()
        path = Path(os.path.abspath(sys.argv[0])).parent.parent
        # metadata
        yfile = path / args.yaml
        with open(yfile) as stream:
            self._yaml = yaml.safe_load(stream)
        # input & output files
        ifile = path / args.input
        odir = path / args.output
        odir.mkdir(parents=True, exist_ok=True)
        ofile = odir / 'catalog.json'
        # initialize helpers
        sheet_name = self._yaml['sheet-name']
        xlsx_helper = XlsxHelper(ifile, sheet_name)
        title = self._yaml['title']
        version = self._yaml['version']
        catalog_helper = CatalogHelper(title, version)
        # process
        self._process(xlsx_helper, catalog_helper)
        # write catalog
        info(f'output: {ofile}')
        catalog = catalog_helper.get_catalog()
        catalog.oscal_write(ofile)

    def _process(self, xlsx_helper: XlsxHelper, catalog_helper: CatalogHelper) -> None:
        """Process."""
        # transform each row into OSCAL equivalent
        for row in xlsx_helper.row_generator():
            group_id = xlsx_helper.get_group_id(row)
            group_title = xlsx_helper.get_group_title(row)
            subgroup_id = xlsx_helper.get_subgroup_id(row)
            subgroup_title = xlsx_helper.get_subgroup_title(row)
            control_id = xlsx_helper.get_control_id(row)
            control_title = xlsx_helper.get_control_title(row)
            control_description = xlsx_helper.get_control_description(row)
            catalog_helper.add_control(
                group_id, group_title, subgroup_id, subgroup_title, control_id, control_title, control_description
            )

    def parse_args(self) -> Dict:
        """Parse args."""
        description = 'Build OSCAL catalog from AI RMF .xlsx file.'
        parser = argparse.ArgumentParser(description=description)
        #
        default_input = 'data/AI-RMF.xlsx'
        help_input = f'.xlsx file containing the AI RMF data; default = {default_input}'
        parser.add_argument(
            '--input',
            default=f'{default_input}',
            action='store',
            help=help_input,
        )
        #
        default_yaml = 'data/AI-RMF.yaml'
        help_yaml = f'.xlsx file containing the AI RMF metadata; default = {default_yaml}'
        parser.add_argument(
            '--yaml',
            default=f'{default_yaml}',
            action='store',
            help=help_yaml,
        )
        #
        default_output = 'catalogs/AI_RMF'
        help_output = f'folder containing the output catalog.json file; default is {default_output}'
        parser.add_argument(
            '--output',
            default=f'{default_output}',
            action='store',
            help=help_output,
        )
        return parser.parse_args()


def main():
    """Mainline."""
    catalog_builder = CatalogBuilder()
    catalog_builder.run()


if __name__ == '__main__':
    main()
